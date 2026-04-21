#!/usr/bin/env python3
"""
extract_boq_rates.py — one-time build step

Parses the three reference BoQ workbooks (Clepington / Lochee / Overton) and
emits src/data/boq_rates_full.js — the authoritative 1,565-item catalogue
consumed by the in-browser BoQ generator.

Usage:
    python3 scripts/extract_boq_rates.py

Requires: openpyxl. Run from the repo root.
"""

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed.  pip install openpyxl", file=sys.stderr)
    sys.exit(1)


REF_XLSX = [
    "Clepington Road - BoQ JMCA.xlsx",
    "Lochee Road- BoQ JMCA.xlsx",
    "Overton-Netherton GRDNS -- BoQ JMCA.xlsx",
]

SERIES_TITLES = {
    100:  "Preliminaries",
    200:  "Site Clearance",
    300:  "Fencing",
    400:  "Safety Barriers",
    500:  "Drainage",
    600:  "Earthworks",
    700:  "Pavements",
    1100: "Footways & Paved Areas",
    1200: "Traffic Signs & Road Markings",
    2700: "Accommodation Works",
    3000: "Landscape & Ecology",
    6400: "Percentage Additions",
}
SERIES_NUMS = list(SERIES_TITLES.keys())

# Sensible default band thresholds per series. Sub-sections within a series
# (e.g. Series 700 carriageway m² vs sub-base m³) have their own thresholds;
# the JS helper picks item-level overrides before falling back to these.
# Empty list means band A is always used (flat-rate series like TM).
SERIES_DEFAULT_THRESHOLDS = {
    100:  [],             # flat-rate preliminaries
    200:  [500, 5000],    # m² site clearance
    300:  [100, 500],     # m fencing
    400:  [100, 500],     # m safety barriers
    500:  [10, 50],       # m drainage runs
    600:  [100, 1000],    # m³ earthworks
    700:  [500, 5000],    # m² pavements (sub-base items override)
    1100: [20, 100],      # m kerbs / m² small footway jobs
    1200: [100, 500],     # m line markings
    2700: [],             # flat per-cover ironwork
    3000: [500, 5000],    # m² landscape
    6400: [],             # percentages — bands carry the different adders
}

# Legacy tag → full-catalogue item id. Drives the boq_rates.js shim so that
# every existing window.boqItem('tag') / window.boqPickRate(...) call keeps
# working after the catalogue swap.
LEGACY_TAG_MAP = {
    # Series 100 — Traffic Management
    "tm_closure_day":      "1/009",
    "tm_diversion_erect":  "1/023",
    "tm_diversion_day":    "1/024",
    "tm_diversion_remove": "1/025",
    "tm_pts_erect":        "1/070",
    "tm_pts_day":          "1/071",
    "tm_pts_remove":       "1/072",
    "tm_sg_erect":         "1/061",
    "tm_sg_day":           "1/062",
    "tm_sg_remove":        "1/063",
    "tm_fw_erect":         "1/098",
    "tm_fw_day":           "1/099",
    "tm_fw_remove":        "1/100",
    # Series 700 — Pavements
    "surf_ac14_40":        "7/027",
    "surf_ac10_40":        "7/029",
    "surf_ac14hb_40":      "7/031",
    "surf_ac10hb_40":      "7/033",
    "surf_hra3014_40_20":  "7/035",
    "surf_hra3014_40_14":  "7/037",
    "surf_hra3514_45_14":  "7/045",
    "surf_hra3514_45_20":  "7/043",
    "surf_sma10_40":       "7/053",
    "surf_sma6_30":        "7/055",
    "surf_hra5510_40":     "7/059",
    "surf_ac6_30":         "7/057",
    "bin_ac20d_60":        "7/019",
    "bin_ac20hdm_60":      "7/021",
    "bin_hra5020_60":      "7/023",
    "bin_ac20cb_60":       "7/025",
    "bin_ac20cb_100":      "7/017",
    "base_ac32d_100":      "7/009",
    "base_ac32d_150":      "7/007",
    "base_ac32d_200":      "7/005",
    "base_ac32hdm_100":    "7/015",
    "base_ac32hdm_150":    "7/013",
    "base_ac32hdm_200":    "7/011",
    "subbase_t1":          "7/001",
    "subbase_rt1":         "7/003",
    "tack":                "7/104",
    "mill_25":             "7/105",
    "mill_40":             "7/106",
    "mill_50":             "7/107",
    "mill_60":             "7/108",
    "mill_70":             "7/109",
    "mill_80":             "7/110",
    "mill_100":            "7/111",
    "mill_150":            "7/112",
    "mill_200":            "7/113",
    "reg_hra3014":         "7/070",
    "reg_ac32d":           "7/063",
    "reg_ac20d":           "7/065",
    "sd_10mm_int":         "7/084",
    "sd_10mm_prem":        "7/085",
    "sd_6mm_int":          "7/086",
    # Series 1100 — Kerbs & Footway
    "kerb_k1_laid":        "11/001",
    "kerb_k1_raised":      "11/002",
    "kerb_k2_laid":        "11/003",
    "kerb_k2_raised":      "11/004",
    "kerb_k3_laid":        "11/005",
    "fw_ac6_20":           "11/194",
    "fw_ac6_30":           "11/195",
    "fw_ac10_30":          "11/196",
    "fw_hra156_20":        "11/197",
    "fw_hra156_30":        "11/198",
    "fw_hra1510_30":       "11/199",
    "fw_subbase_100":      "11/170",
    "fw_subbase_150":      "11/172",
    # Series 1200 — Markings
    "mark_cont_75":        "12/146",
    "mark_cont_100":       "12/149",
    "mark_cont_150":       "12/150",
    "mark_cont_50":        "12/148",
    "mark_int_50_600":     "12/154",
    "mark_int_50_2400":    "12/155",
    "mark_area":           "12/147",
    # Series 2700 — Ironwork
    "iw_sw_cway":          "2700/07",
    "iw_sse_cway":         "2700/23",
    "iw_bt_cway":          "2700/43",
    "iw_sw_fw":            "2700/01",
    "iw_sse_fw":           "2700/13",
    "iw_bt_fw":            "2700/33",
}

ITEM_ID_RE = re.compile(r"^\d+/\d+[A-Z]?$")
NO_OFFER_RE = re.compile(r"^\s*(n/?o|no\s*offer|nil|-+)\s*$", re.IGNORECASE)


def is_no_offer(v):
    if v is None:
        return True
    if isinstance(v, (int, float)):
        return False
    return bool(NO_OFFER_RE.match(str(v)))


def to_num(v):
    """Coerce to float, or return None if not numeric."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 4)
    try:
        return round(float(str(v).replace(",", "").strip()), 4)
    except (ValueError, TypeError):
        return None


def infer_thresholds(row6_vals):
    """Extract up to 2 numeric thresholds from band-label strings like
    '< 50', '50 - 500', '> 500' → [50, 500]."""
    nums = []
    for v in row6_vals:
        if v is None:
            continue
        for match in re.findall(r"(\d+(?:,\d+)?(?:\.\d+)?)", str(v)):
            try:
                n = float(match.replace(",", ""))
                if n not in nums:
                    nums.append(n)
            except ValueError:
                pass
    nums = sorted(set(nums))
    if len(nums) >= 2:
        return [int(nums[0]), int(nums[1])]
    return []


def parse_design_sheet(ws):
    """Return {'thresholds': [...], 'items': [...]} for one Series-N Design sheet.
    Columns: A=item id, B=description, D=unit, E/F/G=Tayside rates (Band A/B/C).

    Uses ws.iter_rows() — in read_only mode, ws.cell(r,c) is O(r) per call so the
    naive nested loop is O(n²). iter_rows streams top-to-bottom once.
    """
    thresholds = []
    items = []

    for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Row 6 carries band threshold labels in columns E–G (or nearby).
        if r == 6:
            thresholds = infer_thresholds(row[4:10] if len(row) >= 10 else row[4:])
            continue

        if len(row) < 7:
            continue
        a = row[0]
        if a is None:
            continue
        a_str = str(a).strip()
        if not ITEM_ID_RE.match(a_str):
            continue

        desc = row[1]
        unit = row[3]
        raw_rates = (row[4], row[5], row[6])
        rA, rB, rC = (to_num(v) for v in raw_rates)

        if all(v is None or is_no_offer(v) for v in raw_rates):
            continue
        if rA is None and rB is None and rC is None:
            continue

        # Fill missing bands so rate-picking never hits None.
        base = rA if rA is not None else (rB if rB is not None else rC)
        rA = rA if rA is not None else base
        rB = rB if rB is not None else base
        rC = rC if rC is not None else base

        items.append({
            "id":    a_str,
            "desc":  " ".join(str(desc).split()) if desc else "",
            "unit":  str(unit).strip() if unit else "Item",
            "rateA": rA,
            "rateB": rB,
            "rateC": rC,
        })

    return {"thresholds": thresholds, "items": items}


def build_catalogue(repo_root):
    catalogue = {}
    seen_ids = {}

    for fname in REF_XLSX:
        path = repo_root / fname
        if not path.exists():
            print(f"  WARN: missing {fname}, skipping", file=sys.stderr)
            continue
        print(f"Reading {fname}")
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            for sn in SERIES_NUMS:
                sheet_name = f"Series {sn} Design"
                if sheet_name not in wb.sheetnames:
                    continue
                parsed = parse_design_sheet(wb[sheet_name])
                key = f"s{sn}"
                if key not in catalogue:
                    catalogue[key] = {
                        "title":          SERIES_TITLES[sn],
                        "bandThresholds": SERIES_DEFAULT_THRESHOLDS.get(sn, []),
                        "items":          [],
                    }

                for it in parsed["items"]:
                    if it["id"] in seen_ids:
                        continue
                    seen_ids[it["id"]] = fname
                    catalogue[key]["items"].append(it)
        finally:
            wb.close()

    return catalogue


HELPERS_JS = r"""
// ── Helpers ──────────────────────────────────────────────────────────────
(function(){
  const F = window.BOQ_RATES_FULL;
  const idIndex = {};
  const tagIndex = {};
  for (const [key, s] of Object.entries(F)) {
    const seriesNum = parseInt(key.slice(1));
    for (const it of s.items) {
      it.series = seriesNum;
      idIndex[it.id] = it;
    }
  }
  for (const [tag, id] of Object.entries(window.BOQ_LEGACY_TAG_MAP)) {
    if (idIndex[id]) tagIndex[tag] = idIndex[id];
  }
  window.BOQ_ID_INDEX  = idIndex;
  window.BOQ_TAG_INDEX = tagIndex;

  // Exact lookup by catalogue id (e.g. '7/027', '2700/43').
  window.boqFullItem = function(id) { return idIndex[id] || null; };

  // Live search across the catalogue. filters: {series:700, unit:'m²'}.
  // Capped at 250 results so huge queries stay snappy.
  window.boqFullSearch = function(q, filters) {
    filters = filters || {};
    const needle = (q || '').toLowerCase().trim();
    const out = [];
    for (const [seriesKey, series] of Object.entries(F)) {
      if (filters.series && 's' + filters.series !== seriesKey) continue;
      for (const it of series.items) {
        if (filters.unit && it.unit !== filters.unit) continue;
        if (needle &&
            !(it.desc || '').toLowerCase().includes(needle) &&
            !(it.id   || '').toLowerCase().includes(needle)) continue;
        out.push(Object.assign({}, it, {
          seriesKey: seriesKey, seriesTitle: series.title,
        }));
        if (out.length >= 250) return out;
      }
    }
    return out;
  };

  // Pick a rate: pass 'A' / 'B' / 'C' to force a band, or a measurement
  // (number) and the series's bandThresholds will decide.
  window.boqPickRateForSeries = function(item, measurement, seriesNum) {
    if (!item) return 0;
    if (measurement === 'A') return +item.rateA || 0;
    if (measurement === 'B') return +item.rateB || 0;
    if (measurement === 'C') return +item.rateC || 0;
    const series = F['s' + (seriesNum || item.series)];
    const thresholds = (item.bandThresholds && item.bandThresholds.length)
      ? item.bandThresholds
      : (series && series.bandThresholds && series.bandThresholds.length
          ? series.bandThresholds : null);
    if (!thresholds || thresholds.length < 2) {
      return +item.rateB || +item.rateA || 0;
    }
    const m = +measurement || 0;
    if (m < thresholds[0]) return +item.rateA || 0;
    if (m < thresholds[1]) return +item.rateB || 0;
    return +item.rateC || 0;
  };

  // Return the band letter ('A'|'B'|'C') that would be picked for a given
  // measurement. Useful for rendering band badges without recomputing rates.
  window.boqPickBand = function(item, measurement, seriesNum) {
    if (measurement === 'A' || measurement === 'B' || measurement === 'C') return measurement;
    const series = F['s' + (seriesNum || (item && item.series))];
    const thresholds = (item && item.bandThresholds && item.bandThresholds.length)
      ? item.bandThresholds
      : (series && series.bandThresholds && series.bandThresholds.length
          ? series.bandThresholds : null);
    if (!thresholds || thresholds.length < 2) return 'B';
    const m = +measurement || 0;
    if (m < thresholds[0]) return 'A';
    if (m < thresholds[1]) return 'B';
    return 'C';
  };
})();
"""


def emit_js(catalogue, out_path):
    lines = []
    lines.append("// boq_rates_full.js — GENERATED by scripts/extract_boq_rates.py")
    lines.append("// Source: Tayside Contracts JMCA rates embedded in the reference")
    lines.append("// Clepington / Lochee / Overton-Netherton BoQ workbooks.")
    lines.append("// DO NOT EDIT BY HAND — rerun the extractor.")
    lines.append("")
    lines.append("window.BOQ_RATES_FULL = {")
    for key in [f"s{n}" for n in SERIES_NUMS]:
        if key not in catalogue:
            continue
        s = catalogue[key]
        lines.append(f"  {key}: {{")
        lines.append(f"    title: {json.dumps(s['title'])},")
        lines.append(f"    bandThresholds: {json.dumps(s['bandThresholds'])},")
        lines.append(f"    items: [")
        for it in s["items"]:
            lines.append("      " + json.dumps(it, ensure_ascii=False) + ",")
        lines.append("    ],")
        lines.append("  },")
    lines.append("};")
    lines.append("")
    lines.append(f"window.BOQ_SERIES_ORDER = {json.dumps(SERIES_NUMS)};")
    lines.append("")
    lines.append("window.BOQ_LEGACY_TAG_MAP = " + json.dumps(LEGACY_TAG_MAP, indent=2) + ";")
    lines.append("")
    lines.append(HELPERS_JS)
    out_path.write_text("\n".join(lines), encoding="utf-8")


def main():
    repo_root = Path(__file__).resolve().parent.parent
    catalogue = build_catalogue(repo_root)

    total = sum(len(s["items"]) for s in catalogue.values())
    print(f"\nTotal items: {total}")
    for key, s in catalogue.items():
        print(f"  {key:>6}: {len(s['items']):>4} items, thresholds={s['bandThresholds']}")

    out_path = repo_root / "src" / "data" / "boq_rates_full.js"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    emit_js(catalogue, out_path)
    size_kb = out_path.stat().st_size / 1024
    print(f"\nWrote {out_path.relative_to(repo_root)} ({size_kb:.1f} KB)")

    # Legacy-tag coverage check
    missing = [tag for tag, iid in LEGACY_TAG_MAP.items()
               if not any(it["id"] == iid for s in catalogue.values() for it in s["items"])]
    if missing:
        print(f"\nWARN: {len(missing)} legacy tags have no catalogue match:", file=sys.stderr)
        for t in missing[:20]:
            print(f"  {t} → {LEGACY_TAG_MAP[t]}", file=sys.stderr)
    else:
        print(f"\nAll {len(LEGACY_TAG_MAP)} legacy tags resolve to catalogue items.")


if __name__ == "__main__":
    main()
